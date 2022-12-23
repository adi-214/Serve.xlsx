import openpyxl
import matplotlib.pyplot as plt
import pandas as pd
import pylab as pl
import numpy as np
from statistics import mean

#=============================================================================================

def isfloat(num):
    try:
        float(num)
        return True
    except ValueError:
        return False

#=============================================================================================

def best_fit_slope_and_intercept(xs, ys):
    m = (((mean(xs) * mean(ys)) - mean(xs * ys)) /
         ((mean(xs) * mean(xs)) - mean(xs * xs)))

    b = mean(ys) - m * mean(xs)

    return m, b

#=======================================================================================

path = "Serve.xlsx"
wb = openpyxl.load_workbook(path)
sheet = wb.active
max_row = sheet.max_row
max_column = sheet.max_column

#============================================================================================

male_tip_amount = {}
for i in range(2, max_row + 1):
    if sheet["A" + str(i)].value == None:
        continue
    if sheet["B" + str(i)].value == None:
        continue
    if sheet["C" + str(i)].value == "Male":
        if isfloat(sheet["A" + str(i)].value) and isfloat(sheet["B" + str(i)].value):
            male_tip_amount[sheet["A" + str(i)].value] = sheet["B" + str(i)].value

male_amount = male_tip_amount.keys()
male_amount = sorted(male_amount)
male_tip = []
for i in male_amount:
    male_tip.append(male_tip_amount[i])


q1 = np.percentile(male_amount, 25)
q3 = np.percentile(male_amount, 75)
iqr = q3 - q1
lower_bound = q1 - (1.5 * iqr)
upper_bound = q3 + (1.5 * iqr)
to_pop_1 = []
for i in range(len(male_amount)):
    if male_amount[i] < lower_bound or male_amount[i] > upper_bound:
        to_pop_1.append(i)

q1 = np.percentile(male_tip, 25)
q3 = np.percentile(male_tip, 75)
iqr = q3 - q1
lower_bound = q1 - (1.5 * iqr)
upper_bound = q3 + (1.5 * iqr)
to_pop_2 = []
for i in range(len(male_tip)):
    if male_tip[i] < lower_bound or male_tip[i] > upper_bound:
        to_pop_2.append(i)

male_amount_x = np.array(male_amount)
male_tip_y = np.array(male_tip)

final_x_1 = np.delete(male_amount_x, to_pop_1)
final_y_1 = np.delete(male_tip_y, to_pop_2)

final_y_1 = np.delete(final_y_1, [final_y_1.size - 1, final_y_1.size - 2])

m1, c1 = best_fit_slope_and_intercept(final_x_1, final_y_1)
regression_line = [(m1*x)+c1 for x in final_x_1]
plt.scatter(final_x_1, final_y_1, color = 'black')
plt.plot(final_x_1, regression_line, color = 'black')

#==============================================================================================

female_tip_amount = {}
for i in range(2, max_row + 1):
    if sheet["A" + str(i)].value == None:
        continue
    if sheet["B" + str(i)].value == None:
        continue
    if sheet["C" + str(i)].value == "Female":
        if isfloat(sheet["A" + str(i)].value) and isfloat(sheet["B" + str(i)].value):
            female_tip_amount[sheet["A" + str(i)].value] = sheet["B" + str(i)].value

female_amount = female_tip_amount.keys()
female_amount = sorted(female_amount)
female_tip = []
for i in female_amount:
    female_tip.append(female_tip_amount[i])


q1 = np.percentile(female_amount, 25)
q3 = np.percentile(female_amount, 75)
iqr = q3 - q1
lower_bound = q1 - (1.5 * iqr)
upper_bound = q3 + (1.5 * iqr)
to_pop_1 = []
for i in range(len(female_amount)):
    if female_amount[i] < lower_bound or female_amount[i] > upper_bound:
        to_pop_1.append(i)

q1 = np.percentile(female_tip, 25)
q3 = np.percentile(female_tip, 75)
iqr = q3 - q1
lower_bound = q1 - (1.5 * iqr)
upper_bound = q3 + (1.5 * iqr)
to_pop_2 = []
for i in range(len(female_tip)):
    if female_tip[i] < lower_bound or female_tip[i] > upper_bound:
        to_pop_2.append(i)

female_amount_x = np.array(female_amount)
female_tip_y = np.array(female_tip)

final_x_1 = np.delete(female_amount_x, to_pop_1)
final_y_1 = np.delete(female_tip_y, to_pop_2)

final_y_1 = np.delete(final_y_1, [final_y_1.size - 1])

m1, c1 = best_fit_slope_and_intercept(final_x_1, final_y_1)
regression_line = [(m1*x)+c1 for x in final_x_1]
plt.scatter(final_x_1, final_y_1, color = 'red')
plt.plot(final_x_1, regression_line, color = 'red')
plt.title("Ammount vs Tip")
plt.legend(["Male","", "Female", ""], loc = "upper right")
plt.show()

#==============================================================================================
lunch_tip_amount = {}
for i in range(2, max_row + 1):
    if sheet["A" + str(i)].value == None:
        continue
    if sheet["B" + str(i)].value == None:
        continue
    if sheet["F" + str(i)].value == "Lunch":
        if isfloat(sheet["A" + str(i)].value) and isfloat(sheet["B" + str(i)].value):
            lunch_tip_amount[sheet["A" + str(i)].value] = sheet["B" + str(i)].value

lunch_amount = lunch_tip_amount.keys()
lunch_amount = sorted(lunch_amount)
lunch_tip = []
for i in lunch_amount:
    lunch_tip.append(lunch_tip_amount[i])


q1 = np.percentile(lunch_amount, 25)
q3 = np.percentile(lunch_amount, 75)
iqr = q3 - q1
lower_bound = q1 - (1.5 * iqr)
upper_bound = q3 + (1.5 * iqr)
to_pop_1 = []
for i in range(len(lunch_amount)):
    if lunch_amount[i] < lower_bound or lunch_amount[i] > upper_bound:
        to_pop_1.append(i)


q1 = np.percentile(lunch_tip, 25)
q3 = np.percentile(lunch_tip, 75)
iqr = q3 - q1
lower_bound = q1 - (1.5 * iqr)
upper_bound = q3 + (1.5 * iqr)
to_pop_2 = []
for i in range(len(lunch_tip)):
    if lunch_tip[i] < lower_bound or lunch_tip[i] > upper_bound:
        to_pop_2.append(i)

lunch_amount_x = np.array(lunch_amount)
lunch_tip_y = np.array(lunch_tip)

lunch_x_1 = np.delete(lunch_amount_x, to_pop_1)
lunch_y_1 = np.delete(lunch_tip_y, to_pop_2)

len_dif_1 = len(lunch_x_1) - len(lunch_y_1)
print(len_dif_1)

m3, c3 = best_fit_slope_and_intercept(lunch_x_1, lunch_y_1)
regression_line = [(m3*x)+c3 for x in lunch_x_1]
plt.scatter(lunch_x_1, lunch_y_1, color = 'black')
plt.plot(lunch_x_1, regression_line, color = 'black')

#==============================================================================================

Dinner_tip_amount = {}
for i in range(2, max_row + 1):
    if sheet["A" + str(i)].value == None:
        continue
    if sheet["B" + str(i)].value == None:
        continue
    if sheet["F" + str(i)].value == "Dinner":
        if isfloat(sheet["A" + str(i)].value) and isfloat(sheet["B" + str(i)].value):
            Dinner_tip_amount[sheet["A" + str(i)].value] = sheet["B" + str(i)].value

Dinner_amount = Dinner_tip_amount.keys()
Dinner_amount = sorted(Dinner_amount)
Dinner_tip = []
for i in Dinner_amount:
    Dinner_tip.append(Dinner_tip_amount[i])


q1 = np.percentile(Dinner_amount, 25)
q3 = np.percentile(Dinner_amount, 75)
iqr = q3 - q1
lower_bound = q1 - (1.5 * iqr)
upper_bound = q3 + (1.5 * iqr)
to_pop_1 = []
for i in range(len(Dinner_amount)):
    if Dinner_amount[i] < lower_bound or Dinner_amount[i] > upper_bound:
        to_pop_1.append(i)


q1 = np.percentile(Dinner_tip, 25)
q3 = np.percentile(Dinner_tip, 75)
iqr = q3 - q1
lower_bound = q1 - (1.5 * iqr)
upper_bound = q3 + (1.5 * iqr)
to_pop_2 = []
for i in range(len(Dinner_tip)):
    if Dinner_tip[i] < lower_bound or Dinner_tip[i] > upper_bound:
        to_pop_2.append(i)

Dinner_amount_x = np.array(Dinner_amount)
Dinner_tip_y = np.array(Dinner_tip)

Dinner_x_1 = np.delete(Dinner_amount_x, to_pop_1)
Dinner_y_1 = np.delete(Dinner_tip_y, to_pop_2)

len_dif_1 = len(Dinner_x_1) - len(Dinner_y_1)
Dinner_y_1 = np.delete(Dinner_y_1, [Dinner_y_1.size - 1, Dinner_y_1.size - 2])

m4, c4 = best_fit_slope_and_intercept(Dinner_x_1, Dinner_y_1)
regression_line = [(m4*x)+c4 for x in Dinner_x_1]
plt.scatter(Dinner_x_1, Dinner_y_1, color = 'red')
plt.plot(Dinner_x_1, regression_line, color = 'red')
plt.title("Ammount vs Tip")
plt.legend(["Lunch","", "Dinner", ""], loc = "upper right")
plt.show()
#==============================================================================================

amount_sun=[]
amount_sat=[]
amount_fri=[]
amount_thurs=[]

for i in range(2, max_row + 1):
    if sheet["A" + str(i)].value == None:
        continue

    if sheet["E" + str(i)].value == "Sun":
        if isfloat(sheet["A" + str(i)].value):
            amount_sun.append(sheet["A" + str(i)].value)

    elif sheet["E" + str(i)].value == "Sat":
        if isfloat(sheet["A" + str(i)].value):
            amount_sat.append(sheet["A" + str(i)].value)

    elif sheet["E" + str(i)].value == "Fri":
        if isfloat(sheet["A" + str(i)].value):
            amount_fri.append(sheet["A" + str(i)].value)

    elif sheet["E" + str(i)].value == "Thurs":
        if isfloat(sheet["A" + str(i)].value):
            amount_thurs.append(sheet["A" + str(i)].value)

amount_sun_new = np.array(amount_sun)
amount_sat_new = np.array(amount_sat)
amount_fri_new = np.array(amount_fri)
amount_thurs_new = np.array(amount_thurs)

chart = np.array([mean(amount_sun_new), mean(amount_sat_new), mean(amount_fri_new), mean(amount_thurs_new)])
days = ["Sun", "Sat", "Fri", "Thurs"]
plt.bar(days, chart, color = 'blue')
plt.title("Bill Amount on Different days")
plt.show()
plt.pie(chart, labels = days, autopct = '%1.1f%%')
plt.title("Bill Amount on Different days")
plt.show()

#=====================================================================================================================

amount_lunch=[]
amount_dinner=[]

for i in range(2, max_row + 1):
    if sheet["A" + str(i)].value == None:
        continue

    if sheet["F" + str(i)].value == "Lunch":
        if isfloat(sheet["A" + str(i)].value):
            amount_lunch.append(sheet["A" + str(i)].value)

    elif sheet["F" + str(i)].value == "Dinner":
        if isfloat(sheet["A" + str(i)].value):
            amount_dinner.append(sheet["A" + str(i)].value)

amount_lunch_new = np.array(amount_lunch)
amount_dinner_new = np.array(amount_dinner)

chart = np.array([mean(amount_lunch_new), mean(amount_dinner_new)])
days = ["Lunch", "Dinner"]
plt.bar(days, chart, color = 'blue')
plt.title("Bill Amount for different meals")
plt.show()
plt.pie(chart, labels = days, autopct = '%1.1f%%')
plt.title("Bill Amount for different meals")
plt.show()

#=======================================================================================================================

party_6 = []
party_5 = []
party_4 = []
party_3 = []
party_2 = []
party_1 = []

for i in range(2, max_row + 1):
    if sheet["A" + str(i)].value == None:
        continue

    if sheet["G" + str(i)].value == 6:
        if isfloat(sheet["A" + str(i)].value):
            party_6.append(sheet["A" + str(i)].value)

    elif sheet["G" + str(i)].value == 5:
        if isfloat(sheet["A" + str(i)].value):
            party_5.append(sheet["A" + str(i)].value)

    elif sheet["G" + str(i)].value == 4:
        if isfloat(sheet["A" + str(i)].value):
            party_4.append(sheet["A" + str(i)].value)

    elif sheet["G" + str(i)].value == 3:
        if isfloat(sheet["A" + str(i)].value):
            party_3.append(sheet["A" + str(i)].value)

    elif sheet["G" + str(i)].value == 2:
        if isfloat(sheet["A" + str(i)].value):
            party_2.append(sheet["A" + str(i)].value)

    elif sheet["G" + str(i)].value == 1:
        if isfloat(sheet["A" + str(i)].value):
            party_1.append(sheet["A" + str(i)].value)

party_6_new = np.array(party_6)
party_5_new = np.array(party_5)
party_4_new = np.array(party_4)
party_3_new = np.array(party_3)
party_2_new = np.array(party_2)
party_1_new = np.array(party_1)

chart_2= np.array([mean(party_6_new), mean(party_5_new), mean(party_4_new), mean(party_3_new), mean(party_2_new), mean(party_1_new)])
days = ["6", "5", "4", "3", "2", "1"]
plt.bar(days, chart_2, color = 'purple')
plt.title("Average bill Amount for different party sizes")
plt.show()
plt.pie(chart_2, labels = days, autopct = '%1.1f%%')
plt.title("Average bill Amount for different party sizes")
plt.show()

#======================================================================================================================

party_6 = []
party_5 = []
party_4 = []
party_3 = []
party_2 = []
party_1 = []

for i in range(2, max_row + 1):
    if sheet["A" + str(i)].value == None:
        continue

    if sheet["G" + str(i)].value == 6:
        if isfloat(sheet["A" + str(i)].value):
            party_6.append(sheet["A" + str(i)].value / 6)

    elif sheet["G" + str(i)].value == 5:
        if isfloat(sheet["A" + str(i)].value):
            party_5.append(sheet["A" + str(i)].value / 5)

    elif sheet["G" + str(i)].value == 4:
        if isfloat(sheet["A" + str(i)].value):
            party_4.append(sheet["A" + str(i)].value / 4)

    elif sheet["G" + str(i)].value == 3:
        if isfloat(sheet["A" + str(i)].value):
            party_3.append(sheet["A" + str(i)].value / 3)

    elif sheet["G" + str(i)].value == 2:
        if isfloat(sheet["A" + str(i)].value):
            party_2.append(sheet["A" + str(i)].value / 2)

    elif sheet["G" + str(i)].value == 1:
        if isfloat(sheet["A" + str(i)].value):
            party_1.append(sheet["A" + str(i)].value)

party_6_new = np.array(party_6)
party_5_new = np.array(party_5)
party_4_new = np.array(party_4)
party_3_new = np.array(party_3)
party_2_new = np.array(party_2)
party_1_new = np.array(party_1)

chart = np.array([mean(party_6_new), mean(party_5_new), mean(party_4_new), mean(party_3_new), mean(party_2_new), mean(party_1_new)])
days = ["6", "5", "4", "3", "2", "1"]
plt.bar(days, chart, color = 'brown')
plt.title("Average bill amount per person for different party sizes")
plt.show()
plt.pie(chart, labels = days, autopct = '%1.1f%%')
plt.title("Average bill amount per person for different party sizes")
plt.show()

#======================================================================================================================

party_sun = []
party_sat = []
party_fri = []
party_thurs = []

for i in range(2, max_row + 1):
    if sheet["G" + str(i)].value == None:
        continue
    if sheet["E" + str(i)].value == "Sun":
        if isfloat(sheet["G" + str(i)].value) and sheet["G" + str(i)].value < 7:
            party_sun.append(sheet["G" + str(i)].value)

    elif sheet["E" + str(i)].value == "Sat":
        if isfloat(sheet["G" + str(i)].value) and sheet["G" + str(i)].value < 7:
            party_sat.append(sheet["G" + str(i)].value)

    elif sheet["E" + str(i)].value == "Fri":
        if isfloat(sheet["G" + str(i)].value) and sheet["G" + str(i)].value < 7:
            party_fri.append(sheet["G" + str(i)].value)

    elif sheet["E" + str(i)].value == "Thurs":
        if isfloat(sheet["G" + str(i)].value) and sheet["G" + str(i)].value < 7:
            party_thurs.append(sheet["G" + str(i)].value)

party_sun_new = np.array(party_sun)
party_sat_new = np.array(party_sat)
party_fri_new = np.array(party_fri)
party_thurs_new = np.array(party_thurs)

chart_3 = np.array([mean(party_sun_new), mean(party_sat_new), mean(party_fri_new), mean(party_thurs_new)])
days = ["Sunday", "Saturday", "Friday", "Thursday"]
plt.bar(days, chart_3, color = 'green')
plt.title("Average party size for different days")
plt.show()
plt.pie(chart_3, labels = days, autopct = '%1.1f%%')
plt.title("Average party size for different days")
plt.show()

#======================================================================================================================

party_lunch = []
party_dinner = []

for i in range(2, max_row + 1):
    if sheet["G" + str(i)].value == None:
        continue

    if sheet["F" + str(i)].value == "Lunch":
        if isfloat(sheet["G" + str(i)].value) and sheet["G" + str(i)].value < 7:
            party_lunch.append(sheet["G" + str(i)].value)

    if sheet["F" + str(i)].value == "Dinner":
        if isfloat(sheet["G" + str(i)].value) and sheet["G" + str(i)].value < 7:
            party_dinner.append(sheet["G" + str(i)].value)

party_lunch_new = np.array(party_lunch)
party_dinner_new = np.array(party_dinner)

chart = np.array([mean(party_lunch_new), mean(party_dinner_new)])
days = ["Lunch", "Dinner"]

plt.bar(days, chart, color = 'red')
plt.title("Average party size for different meals")
plt.show()
plt.pie(chart, labels = days, autopct = '%1.1f%%')
plt.title("Average party size for different meals")
plt.show()

#=====================================================================================================================

amount_smoker = []
amount_non_smoker = []

for i in range(2, max_row + 1):
    if sheet["A" + str(i)].value == None:
        continue

    if sheet["D" + str(i)].value == "Yes":
        if isfloat(sheet["A" + str(i)].value):
            amount_smoker.append(sheet["A" + str(i)].value)

    elif sheet["D" + str(i)].value == "No":
        if isfloat(sheet["A" + str(i)].value):
            amount_non_smoker.append(sheet["A" + str(i)].value)

amount_smoker_new = np.array(amount_smoker)
amount_non_smoker_new = np.array(amount_non_smoker)

chart = np.array([mean(amount_smoker_new), mean(amount_non_smoker_new)])
days = ["Smoker", "Non-Smoker"]

plt.bar(days, chart, color = 'blue')
plt.title("Average bill amount for smokers vs non smokers")
plt.show()
plt.pie(chart, labels = days, autopct = '%1.1f%%')
plt.title("Average bill amount for smokers vs non smokers")
plt.show()

#=====================================================================================================================